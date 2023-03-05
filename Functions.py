import openpyxl
import re
import PySimpleGUI as pg
import matplotlib


matplotlib.use('Agg')
Verficador={}
Verficando={}
diff1={}
diff={}
doppel={}

def Begin():                                              #Função principal, começa o programa e chama todas as outas
    pg.theme("DarkBlue")
    layout=[                                                                                        #Cria a tela incial
              [pg.Text('Lista de Equipamentos:', size=(13,2))],
              [pg.Input(), pg.FileBrowse(key='-Lista-',file_types=(("Excel", "*.xlsx"),("Excel", "*.xlsm")))],
              [pg.Text('Planilha de Quantidades: ', size=(13,2))],
              [pg.Input(), pg.FileBrowse(key='-Pq-',file_types=(("Excel", "*.xlsx"),("Excel", "*.xlsm")))],
              [pg.Button("Submit"), pg.Button("Cancel")]
          ]
    window= pg.Window('Verificador',layout, size=(450,200))
    event,values = window.read()
    window.close()
    open_Le(values['-Lista-'])                                         # abre a Le e faz todas as operações necessárias
    open_PQ(values['-Pq-'])                                             #abre a PQ e faz todas as operações necessárias
    check_PQ_LE()                                                      #Checa as incosistências PQ -> LE
    check_LE_PQ()                                                        #Checa as incosistências LE -> PQ
    open_popup()                                                         #Feito as operações de comparação, é aberto a janela de inconsistências

   
def check_pattern(letter):                                                # Verfica se o conteúdo da célula é um tag
  pattern= '^(\D{2,3})-(\w{5,10})-(\d{2,4})$'                              # \D{2,3}= entre 2 a 3 não dígitos -(\w{5,10})- = -entre 5 a 10 alfanúmericos-  \d{2,4} = entre 2 a 4 dígitos
  match = re.search(pattern, letter)
  if (match != None):
     return True
  else:
     return False      
 
def open_Le(element):                                                     # abre a LE e realiza as operações de iterar, testar valores, formatar e guardar o conteúdo em um dicionário
  aux=str()
  theFile = openpyxl.load_workbook(element)
  allSheetNames = theFile.sheetnames
  for sheet in allSheetNames:
     currentSheet = theFile[sheet]
     for row in range(1, currentSheet.max_row + 1):
          cell_name = "{}{}".format("A", row)
          cell_area= "{}{}".format("I", row)
          cell_value= "{}{}".format("G", row)
          if (currentSheet[cell_name].value == "-"):
            aux1=untagged_Le(currentSheet[cell_area].value)
            Verficando[f'Sem tagg {aux1}'] =  Verficando.get(f'Sem tagg {aux1}',0) + currentSheet[cell_value].value
          else:
            if (currentSheet[cell_name].value != None and check_pattern(str(currentSheet[cell_name].value)) == True):
              aux=formato(currentSheet[cell_name].value)
              if(check_duplicates_Le(aux,cell_name)):
                Verficando[aux]= f'Equipamentos:{cell_name}'


  return Verficando

def untagged_Le(letter):
   indice = re.finditer(pattern=' ',string=letter)
   ind=[index.start() for index in indice]
   return letter[:ind[0]]

def open_PQ(element):
  exit=0
  fm_list=[]                                               # abre a PQ e realiza as operações de iterar, testar valores, formatar e guardar o conteúdo em um dicionário
  theFile = openpyxl.load_workbook(element)
  allSheetNames = theFile.sheetnames
  for sheet in allSheetNames:
      currentSheet = theFile[sheet]
      for row in range(12, currentSheet.max_row + 1):
          cell_name = "{}{}".format("D", row)
          if (currentSheet[cell_name].value != None):
            cell_auxc = "{}{}".format("C", row)
            cell_auxb = "{}{}".format("B", row)
            cell_value = "{}{}".format("Q", row)
            cell_fm = "{}{}".format("E", row)
            if(diff_FM(fm_list, str(currentSheet[cell_fm].value)) and currentSheet[cell_fm].value != "E"):
              if(currentSheet[cell_name].value == "-"):
                Verficador[f'Sem tagg {str(currentSheet[cell_auxb].value) + str(currentSheet[cell_auxc].value)}'] =  Verficador.get(f'Sem tagg {str(currentSheet[cell_auxb].value) + str(currentSheet[cell_auxc].value)}',0) + currentSheet[cell_value].value
              else:
                  aux=currentSheet[cell_name].value[:3] + str(currentSheet[cell_auxb].value) + str(currentSheet[cell_auxc].value) +"-"+ currentSheet[cell_name].value[3:]
                  if (check_pattern(aux)): 
                    aux=formato(aux)
                    if(check_duplicates_Pq(aux,cell_name)):
                      Verficador[aux]= f'Sheet:{sheet}:{cell_name}'
            else:
              break
          
  return Verficador  

def formato(let):                                                # formata os tags econtrados visto que podem haver 1 ou 2 zeros antes do número de fato, ex : PB-6027SA-04 ou IT-6027SA-008
  new_string=str()
  indice = re.finditer(pattern='-',string=let)
  ind=[index.start() for index in indice]
  if ( let[ind[1]+1] == '0' and let[ind[1]+2] != '0' ):
    new_string = let[:ind[1]+1] + let[ind[1]+2:]
  elif(let[ind[1]+1] == '0' and let[ind[1]+2] == '0'):
     new_string = let[:ind[1]+1] + let[ind[1]+3:]
  else:
     new_string=let
  return new_string
    
def check_PQ_LE():
  for element, val in Verficador.items():
    if ( not (element in Verficando)):
      diff[element]=val
    ##else:
      ## if(element.find("Sem tagg") != -1):
        ##  diff[element]= val - diff1[element]
 
def check_LE_PQ():
    for element, val in Verficando.items():
      if ( (not (element in Verficador.keys()))):
        diff1[element]=val
      ##else:
        ##if(element.find("Sem tagg") != -1):
          ##  diff1[element]= val - diff[element]
         

def check_duplicates_Le(aux,cell):
   if (aux in Verficando):
      doppel[aux]=f'Le:{cell}'
      return False
   else:
      return True
   
def check_duplicates_Pq(aux,cell):
   if (aux in Verficador):
      doppel[aux]= f'PQ:{cell}'
      return False
   else:
      return True
   
def diff_FM(list,string):
  if(string=="F" or string == "M"):
      if(len(list)==0):
        list.append(string)
        return True
      
      else:
         for element in list:
            if(string==element):
              return True
            else:
              return False
  else:
     return True
            

def open_popup():
  if(len(diff)==0 and len(diff1) == 0 and len(doppel)==0):
      pg.theme("DarkAmber")  
      layout=[
          [pg.Text("Nenhuma Iconsistência encontrada")],
          [pg.Button("OK"), pg.Button("Cancel")]
      ]
      window= pg.Window("Tabela de Inconsistências",layout)
      while True:
          print(window.read())
          break
  else:
      pg.theme("DarkRed")
      s = '\n***Itens não presentes na Lista***\n\n'
      s += '\n'.join([str(i) for i in diff.items()])
      s1 = '\n***Itens não presentes na PQ***\n\n'
      s1 += '\n'.join([str(i) for i in diff1.items()])
      s2 =  '\n***Itens Duplicados***\n\n'
      s2 += '\n'.join([str(i) for i in doppel.items()])
      text= s+'\n'+s1+'\n'+s2
      column = [[pg.Text(text, font=('Courier New', 12))]]
      layout=[
          [pg.Text(f'Inconsistências Encontradas:', font=('Bold', 16))],
          [pg.Column(column, size=(800, 300), scrollable=True, key = "Column")],
          [pg.Button("OK"), pg.Button("Cancel")]
      ]
      window= pg.Window("Tabela de Inconsistências",layout)
      while True:
          print(window.read())
          break 
       
   
    

