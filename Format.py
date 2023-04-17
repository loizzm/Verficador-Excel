import openpyxl
from openpyxl.styles import PatternFill
import PySimpleGUI as pg 
import csv
import array
import pickle
import string


def find_color(color,lista):                                            ## Serve para procurar a última vez que uma cor foi encontrada, visto que normalmente existe itens em branco antes das cores principais
      if (color=='FF9900' or color=='00FF9900'):                        ## Nõa há branco descrito nessa função, pois nunca o branco virá entre duas cores diferentes de branco
          indices = [i for i, x in enumerate(lista) if (x.count('.')==0)]     ## retorna os indíces de todas as vezes que essa cor foi encontrada 
          return indices[-1]                                            ## retorna o último índide onde a cor foi encontrada
      elif(color=='FFCC00'or color=='00FFCC00'):
          indices = [i for i, x in enumerate(lista) if ((x.count('.')==1))]
          return indices[-1]
      elif(color == 'FFE68D' or color  == '00FFE68D' ):
          indices = [i for i, x in enumerate(lista) if (x.count('.')==2)]
          return indices[-1]
      elif(color == 'FFFF99' or color == '00FFFF99' ):
          indices = [i for i, x in enumerate(lista) if (x.count('.')==3)]
          return indices[-1]    

def c_by_v(color,lista):                                                 ## Retorna o valor destinado a célula baseado na cor e na posição dela na lista
    if(len(lista)==0):
        if(color=='FF9900' or color=='00FF9900'):                        ## Para a priemeira vez que a cor laranja for encontrada
            lista.append("1")
            return str(lista[-1])
    else:
        if(color=='FF9900' or color=='00FF9900'): 
            value = lista[find_color(color,lista)]                      ## Para a segunda vez que a cor laranja for encontrada, antecipada por uma célula branca
            aux = int(value)                                             ## Transfomra "1" em 1
            aux +=1                                                      ## acrescenta
            string= str(aux)                                             ## Transforma em string dnv
            lista.append(string)                                         ## acrescenta na lista
            return str(lista[-1]) 
        
        elif((color=='FFCC00'or color=='00FFCC00') and (lista[-1].count('.')==0)):   ##  a primeira aparição de uma cor laranja mais amarelada, sempre é antecipada pelo laranja forte por isso o len(lista[-1])==1
            value = float(lista[-1])                                         ## transforma em float o último valor da lista
            value += 0.1                                                     ## Faz 1.0 virar 1.1
            value= round(value,2)                                            ## arredonda para duas casas decimais já que transfomações pra float podem ter erros de arredondamneto                             
            value=str(value)                                                 ## Transforma em string novamente
            lista.append(value)
            return value
        
        elif((color=='FFCC00'or color=='00FFCC00') and  (lista[-1].count('.')==4)):    ## para uma aparição da cor laranja mais amarelado depois de um branco, por isso o len do úlitmo elemneto regisrado na lista tem que ser maior que 9
            value = lista[find_color(color,lista)]
            aux= float(value)                                                ## Transfoma em float
            aux += 0.1                                                       ## soma 0,1 n0 valor anterior, ou seja "1.1 + 0.1 = 1.2"
            aux=round(aux,2)                                                 ## Arredonda para 2 casas decimais 
            string=str(aux)                                                  ## Transforma em string
            lista.append(string)
            return str(lista[-1])
        
        elif((color == 'FFE68D' or color  == '00FFE68D' ) and (lista[-1].count('.')==1)):                        ## Primeira aparição do amarelo, logo depois de um laranja mais amarelad0, por isso o len do úlitmo item da lista tem que ter tamanho 3("1.1")
            value = (lista[-1])
            value = value + '.1'
            lista.append(value)
            return value
        
        elif((color == 'FFE68D' or color  == '00FFE68D' ) and (lista[-1].count('.')==4)):                        ## Para um aparição depois de uma célula branca, por isso o len do úlitmo item da lista tem que ter tamanho maior que 9
            value = lista[find_color(color,lista)]
            index=value.rfind('.')
            aux = int(value[index+1:])                                            ##Pega o último valor da última aparição registrada dessa cor, por isso o 4 ("1.1.1")
            aux +=1
            string= str(aux)
            value= value[:index+1]+string
            lista.append(value)
            return(value)

        elif((color == 'FFFF99' or color == '00FFFF99') and (lista[-1].count('.')==2)):                     ##Para uma primeira aparição do amarelo mais fraco, vindo depois de um amrelo, por isso o len ==5 (1.1.1)
            value = (lista[-1])
            value = value + '.1'
            lista.append(value)
            return value
        
        elif((color == 'FFFF99' or color == '00FFFF99') and (lista[-1].count('.')==4)):                   ## Para uma aparição do amareloa mais fraco depois de um branco, por isso o len > 9
            value= lista[find_color(color,lista)]
            index=value.rfind('.')
            aux = int(value[index+1:])
            aux +=1
            string= str(aux)
            value= value[:index+1]+string
            lista.append(value)
            return(value)
        
        elif((color == '000000' or color == '00FFFFFF')and len(lista) >= 4):                  ## Para uma parição de células Brancas
            if((lista[-1].count('.')==3)):
                value = (lista[-1])
                value = value + '.1'
                lista.append(value)
                return value
            else:
                value = (lista[-1])
                if(lista[-1].count('.')==4):
                    index=value.rfind('.')
                    value=int(value[index+1:])
                    value +=1
                    value = str(value)
                    aux= lista[-1]
                    aux= aux[:index+1] + value
                    lista.append(aux)
                    return aux
       
def list_create(c1,c2,c3,c4):
    count=0
    arr = array.array('i', [0]*4)
    lista=[c1,c2,c3,c4]
    for item in lista:
        if(item.count("0") == len(item)):
            count+=1
    if(count==4):
        arr=([0,0,0,0])
        return arr
    elif(count==3):
        arr=([1,0,0,0])
        return arr
    elif(count==2):
        arr=([1,1,0,0])
        return arr
    elif(count==1):
        arr=([1,1,1,0])
        return arr
    elif(count==0):
        arr=([1,1,1,1])
        return arr
    
def predictions(data):
    with open('modelo.pkl', 'rb') as f:
        modelo = pickle.load(f)
    predicoes = modelo.predict(data)
    return predicoes

def csv_create(lista1,lista2):
    with open("dados.csv", "w", newline="") as arquivo_csv:
        escritor_csv = csv.writer(arquivo_csv)
        escritor_csv.writerow(["Cms", "Cor"])
        for lista1, lista2 in zip(lista1, lista2):
            escritor_csv.writerow([lista1, lista2])


def File_name(string):
    if (string.find(".xlsx")!=-1):
        indice =string.find(".xlsx")
        aux=string[:indice] +"-ITEMIZADO"+ string[indice:]
        return aux

def check_pattern_cms(letter):
  if (letter.isdigit()):
     return True
  else:
     return False      

def row_color(row, color,sheet):
       	for letter in string.ascii_uppercase:
            if (letter=='U' or (color == '000000' or color == '00FFFFFF')):
                break
            cell_name = "{}{}".format(letter, row)
            cell = sheet[cell_name]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.fill= fill
            

def Begin():
  pg.theme("DarkBlue")
  layout=[                                                                                       
              [pg.Text('Planilha de Quantidades: ', size=(13,2))],
              [pg.Input(), pg.FileBrowse(key='-Pq-',file_types=(('Excel', '*.xlsx'),))],
              [pg.Button("Submit"), pg.Button("Cancel")]
   ]
  window= pg.Window('Verificador',layout, size=(450,150))
  event,values = window.read()
  window.close()
  cms=[]
  lista=[]
  colors=[]
  theFile = openpyxl.load_workbook(values['-Pq-'])
  allSheetNames = theFile.sheetnames
  for sheet in allSheetNames:
      begin=False
      cms.clear()
      lista.clear()
      colors.clear()                                    ## Reseta a lista para ser preenchida novamente na próxima sheet
      currentSheet = theFile[sheet]
      print(currentSheet)
      if (currentSheet.sheet_state == 'hidden' or currentSheet.sheet_state == 'veryHidden'):
          continue
      for row in range(1, currentSheet.max_row + 1):
            cell_name = "{}{}".format("A", row)
            cell_exit= "{}{}".format("I", row)
            cell_name1 = "{}{}".format("F", row)
            cell_name2 = "{}{}".format("G", row)
            cell_name3 = "{}{}".format("H", row)
            cell_name4 = "{}{}".format("I", row)
            cell = currentSheet[cell_name]
            cell1 = currentSheet[cell_name1]
            cell2 = currentSheet[cell_name2]
            cell3 = currentSheet[cell_name3]
            cell4 = currentSheet[cell_name4]
            print(cell)
            if (type(cell).__name__ != 'MergedCell'):
                string= str(cell1.value) + str(cell2.value) + str(cell3.value) + str(cell4.value)
                if(check_pattern_cms(string)):    
                    if(currentSheet[cell_exit].value == None and len(lista) != 0):                          ##Para resolver o problema do código escrever onde não devia
                        begin=False
                        break 
                    color_hex=list_create(str(cell1.value), str(cell2.value), str(cell3.value), str(cell4.value))
                    cms.append(color_hex)
                    cor_pred=predictions(cms)
                    if(cor_pred[0]=='00FF9900'):
                        begin=True
                    if(begin==True):
                        row_color(row,cor_pred[-1],currentSheet)
                        cell.value= c_by_v(cor_pred[-1],lista)
                        colors.append(cor_pred[-1])
                        
  #csv_create(cms,colors)
  pg.popup_auto_close("Feito!!")
  theFile.save(File_name(values['-Pq-']))
