import openpyxl
import re
import PySimpleGUI as pg

dicc={}
Verificador={}
dict1={}
Verificando={}
wpot={}
diff={}
diff1={}

def Begin():
      pg.theme("DarkBlue")
      layout=[                                                                                        #Cria a tela incial
              [pg.Text('Tabela de demanda:', size=(13,2))],
              [pg.Input(), pg.FileBrowse(key='-Tabela-',file_types=(("Excel", "*.xlsx"),("Excel", "*.xlsm")))],
              [pg.Text('Lista de Equipamentos: ', size=(13,2))],
              [pg.Input(), pg.FileBrowse(key='-Lista-',file_types=(("Excel", "*.xlsx"),("Excel", "*.xlsm")))],
              [pg.Button("Submit"), pg.Button("Cancel")]
          ]
      window= pg.Window('Verificador',layout, size=(450,200))
      event,values = window.read()
      window.close()
      open_Le(values['-Lista-'])                                         # abre a Le e faz todas as operações necessárias
      open_TD(values['-Tabela-']) 
      check_PQ_LE(Verificando,Verificador)
      check_LE_PQ(Verificador,Verificando)
      open_popup()




def check_PQ_LE(Verficador,Verficando):
  for element, val in Verficador.items():
    if ( not (element in Verficando)):
      diff[element]=val
  return diff

def check_LE_PQ(Verficando,Verficador):
    for element, val in Verficando.items():
      if ( (not (element in Verficador.keys()))):
        diff1[element]=val
      else:
         ind = val.find("Potência:")
         aux= val[ind+9:]
         string=Verficador[element]
         ind1=string.find("Potência:")
         aux1= string[ind1+9:]
         if(aux1.find("(")!=-1):
          indice=aux1.find("(")
          palavra=aux1[:indice]
          if(palavra.strip()!=aux):
              wpot[element]=f'Potências divergentes {aux}/{palavra}'
         else:
          if(aux1!=aux):
              wpot[element]=f'Potências divergentes {aux}/{aux1}'
         
  

def check_pattern(letter):                                                # Verfica se o conteúdo da célula é um tag
  pattern= '^(\D{2,3})-(\w{5,10})-(\d{2,4})$'                              # \D{2,3}= entre 2 a 3 não dígitos -(\w{5,10})- = -entre 5 a 10 alfanúmericos-  \d{2,4} = entre 2 a 4 dígitos
  match = re.search(pattern, letter)
  if (match != None):
     return True
  else:
     return False 
  
def wrong_pattern(dict, string,string1):
   pattern= '^(\D{2,3})-(\w{5,10})-(\d{5})$'
   pattern1= '^(\D{1})-(\w{5,10})-(\d{2,5})$'
   pattern2= '^(\D{2,3})-(\w{5,10})-(\d{1})$'
   pattern3= '^(\D{2,3})-(\w{5,10})-$'
   match = re.search(pattern, string)
   match1 = re.search(pattern1, string)
   match2 = re.search(pattern2, string)
   match3 = re.search(pattern3, string)
   if (match != None or match1 != None or match2 != None or match3 != None):
      dict[string]=string1
      
def formato(let):                                                # formata os tags econtrados visto que podem haver 1 ou 2 zeros antes do número de fato, ex : PB-6027SA-04 ou IT-6027SA-008
  new_string=str()
  indice = re.finditer(pattern='-',string=let)
  ind=[index.start() for index in indice]
  if ( let[ind[1]+1] == '0' and let[ind[1]+2] != '0' ):
    new_string = let[:ind[1]+1] + let[ind[1]+2:]
  elif(let[ind[1]+1] == '0' and let[ind[1]+2] == '0'):
     new_string = let[:ind[1]+1] + let[ind[1]+3:]
  else:
     new_string=let
  return new_string



def open_TD(reference):
  theFile = openpyxl.load_workbook(reference)
  allSheetNames = theFile.sheetnames
  for sheet in allSheetNames:
      currentSheet = theFile[sheet]
      for row in range(1, currentSheet.max_row + 1):
            cell_name = "{}{}".format("C", row)
            cell_pot = "{}{}".format("L", row)
            if (currentSheet[cell_name].value != None):
              wrong_pattern(dicc,str(currentSheet[cell_name].value),str(cell_name))
              if (check_pattern(str(currentSheet[cell_name].value)) == True):
                      aux=formato(currentSheet[cell_name].value)
                      Verificador[aux]= f'Sheet:{sheet}:{cell_name} Potência:{currentSheet[cell_pot].value}'


def open_Le(reference):
  theFile1 = openpyxl.load_workbook(reference)
  allSheetNames = theFile1.sheetnames
  for sheet in allSheetNames:
      currentSheet = theFile1[sheet]
      for row in range(1, currentSheet.max_row + 1):
            cell_name = "{}{}".format("A", row)
            cell_pot = "{}{}".format("O", row)
            if (currentSheet[cell_name].value != None):
              wrong_pattern(dict1,str(currentSheet[cell_name].value),str(cell_name))
              if (check_pattern(str(currentSheet[cell_name].value)) == True):
                      aux=formato(currentSheet[cell_name].value)
                      Verificando[aux]= f'Sheet:{sheet}:{cell_name} Potência:{currentSheet[cell_pot].value}'

                       
def open_popup():
  if(len(dict1)==0 and len(dicc) == 0 and len(diff)==0 and len(wpot)==0 and len(diff1)==0):
      pg.theme("DarkAmber")  
      layout=[
          [pg.Text("Nenhuma Iconsistência encontrada")],
          [pg.Button("OK"), pg.Button("Cancel")]
      ]
      window= pg.Window("Tabela de Inconsistências",layout)
      while True:
          print(window.read())
          break
  else:
      pg.theme("DarkRed")
      s = '\n***Itens não presentes na Tabela de Demandas***\n\n'
      s += '\n'.join([str(i) for i in diff.items()])
      s1 = '\n***Itens não presentes na Lista***\n\n'
      s1 += '\n'.join([str(i) for i in diff1.items()])
      s2 =  '\n***Potências Inconsistentes***\n\n'
      s2 += '\n'.join([str(i) for i in wpot.items()])
      s3 =  '\n***Tags inconsistentes na Tabela de Demanda***\n\n'
      s3 += '\n'.join([str(i) for i in dicc.items()])
      s4 =  '\n***Tags inconsistentes na Lista de Materiais***\n\n'
      s4 += '\n'.join([str(i) for i in dict1.items()])
      text= s+'\n'+s1+'\n'+s2+s3+'\n'+s4
      column = [[pg.Text(text, font=('Courier New', 12))]]
      layout=[
          [pg.Text(f'Inconsistências Encontradas:', font=('Bold', 16))],
          [pg.Column(column, size=(800, 300), scrollable=True, key = "Column")],
          [pg.Button("OK"), pg.Button("Cancel")]
      ]
      window= pg.Window("Tabela de Inconsistências",layout)
      while True:
          print(window.read())
          break 
